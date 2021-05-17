import openpyxl

wb = openpyxl.Workbook()

wb.create_sheet('hoja1')

wb.save('prueba.xlsx')

wb.close()



wb1 = openpyxl.load_workbook('prueba.xlsx')

#A que hoja quiero acceder
hoja = wb1.active

#A la celda que quiero acceder
hoja['A1'] = 'valor'
hoja.cell(row=2,column=2).value = 10

print(hoja['A1'].value)
print(hoja.cell(row=1,column=1).value)

#Acceder al ultimo valor de la columna ingresada
print(hoja.max_column)
print(hoja.max_row)

wb1.save('prueba.xlsx')

wb1.close()

